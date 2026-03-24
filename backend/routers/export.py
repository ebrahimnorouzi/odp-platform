"""Export: CSV and Excel download of all responses for a survey."""
import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from database import get_db, Survey, Response
from auth import get_current_admin

router = APIRouter(prefix="/api/surveys", tags=["export"])


def _rows(survey: Survey, responses: list[Response]):
    qs       = survey.questions
    q_ids    = [q["id"] for q in qs]
    headers  = [
        "survey_id","survey_title","session_num",
        "pattern_id","pattern_title",
        "started_at","completed_at","duration_ms","duration_human",
        "submitted_at","session_duration_ms","session_duration_human",
        *[f"q_{qid}" for qid in q_ids],
    ]

    def ms(v):
        if not v: return ""
        s = v // 1000; m,s = divmod(s,60); h,m = divmod(m,60)
        return f"{h}h {m}m {s}s" if h else (f"{m}m {s}s" if m else f"{s}s")

    rows = []
    for r in responses:
        row = {
            "survey_id":              survey.id,
            "survey_title":           survey.title,
            "session_num":            r.session.num if r.session else "",
            "pattern_id":             r.pattern_id,
            "pattern_title":          r.pattern_title,
            "started_at":             r.started_at.isoformat()  if r.started_at  else "",
            "completed_at":           r.completed_at.isoformat() if r.completed_at else "",
            "duration_ms":            r.duration_ms or "",
            "duration_human":         ms(r.duration_ms),
            "submitted_at":           r.submitted_at.isoformat() if r.submitted_at else "",
            "session_duration_ms":    r.session_duration_ms or "",
            "session_duration_human": ms(r.session_duration_ms),
        }
        for qid in q_ids:
            row[f"q_{qid}"] = r.answers.get(qid, "")
        rows.append(row)
    return headers, rows


@router.get("/{sid}/export/csv")
def export_csv(sid: int, db: Session = Depends(get_db), _=Depends(get_current_admin)):
    survey    = db.get(Survey, sid)
    if not survey: raise HTTPException(404)
    responses = db.query(Response).filter(Response.survey_id == sid).order_by(Response.submitted_at).all()
    headers, rows = _rows(survey, responses)

    def esc(v):
        s = str(v)
        return f'"{s.replace(chr(34), chr(34)*2)}"' if any(c in s for c in ',"\n') else s

    buf = "\ufeff" + ",".join(f'"{h}"' for h in headers) + "\n"
    for row in rows:
        buf += ",".join(esc(row.get(h,"")) for h in headers) + "\n"

    return StreamingResponse(
        iter([buf]), media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="survey-{sid}-responses.csv"'},
    )


@router.get("/{sid}/export/excel")
def export_excel(sid: int, db: Session = Depends(get_db), _=Depends(get_current_admin)):
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    survey    = db.get(Survey, sid)
    if not survey: raise HTTPException(404)
    responses = db.query(Response).filter(Response.survey_id == sid).order_by(Response.submitted_at).all()
    headers, rows = _rows(survey, responses)

    df = pd.DataFrame(rows, columns=headers)
    qs = survey.questions
    summary = []
    for q in qs:
        col  = f"q_{q['id']}"
        vals = pd.to_numeric(df[col], errors="coerce").dropna() if col in df else pd.Series([], dtype=float)
        summary.append({
            "ID": q["id"], "Label": q.get("label",""), "Type": q.get("type",""),
            "N":   len(vals) if q.get("type")=="likert" else (df[col].notna().sum() if col in df else 0),
            "Mean": round(vals.mean(),3) if len(vals)>0 and q.get("type")=="likert" else "",
            "Std":  round(vals.std(),3)  if len(vals)>1 and q.get("type")=="likert" else "",
            "Min":  vals.min() if len(vals)>0 and q.get("type")=="likert" else "",
            "Max":  vals.max() if len(vals)>0 and q.get("type")=="likert" else "",
        })
    df_sum = pd.DataFrame(summary)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All Responses", index=False)
        df_sum.to_excel(writer, sheet_name="Question Summary", index=False)
        ws = writer.sheets["All Responses"]
        for cell in ws[1]:
            cell.font  = Font(bold=True, color="FFFFFF")
            cell.fill  = PatternFill("solid", fgColor="0B1022")
            cell.alignment = Alignment(horizontal="center")
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, min(40, len(h)+4))
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="survey-{sid}.xlsx"'},
    )
